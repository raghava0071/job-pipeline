#!/opt/anaconda3/bin/python3
# =============================================================================
# MASTER_RUN.PY  v8  — PER-JOB SEQUENTIAL PIPELINE
# Powered by Claude claude-sonnet-4-6
#
# FLOW (for EVERY job, one at a time):
#   1. Load job from linkedin_jobs.csv / indeed_jobs.csv
#   2. Level filter  → skip Senior/Lead/Director/Manager
#   3. Claude score  → skip if < 65%
#   4. Parse JD      → extract ATS keywords + before/after score
#   5. Build resume  → save .docx to ~/job_pipeline/resumes/
#   6. Cover letter  → save .docx to ~/job_pipeline/cover_letters/
#   7. Apply         → LinkedIn Easy Apply OR Indeed In-Portal (browser)
#   8. Log result    → apply_log.json + Application_Tracker.xlsx
#   9. Next job
#
# USAGE:
#   python master_run.py                    # score + resume + cover letter + apply
#   python master_run.py --dry-run          # score + resume + cover letter, NO apply
#   python master_run.py --skip-fetch       # use existing CSVs (default behaviour)
#   python master_run.py --limit 5          # only process 5 jobs
#   python master_run.py --source linkedin  # only LinkedIn jobs
#   python master_run.py --source indeed    # only Indeed jobs
#   python master_run.py --no-apply         # build resumes only, don't open browser
# =============================================================================

import sys
import time
import json
import argparse
from pathlib import Path
from datetime import datetime

# ── Always operate on ~/job_pipeline regardless of where script is run from ───
PIPELINE_DIR = Path.home() / "job_pipeline"
DATA_DIR     = PIPELINE_DIR / "data"
RESUMES_DIR  = PIPELINE_DIR / "resumes"
CL_DIR       = PIPELINE_DIR / "cover_letters"
APPLY_LOG    = DATA_DIR / "apply_log.json"
SCREENSHOTS  = DATA_DIR / "screenshots"

for d in [DATA_DIR, RESUMES_DIR, CL_DIR, SCREENSHOTS]:
    d.mkdir(parents=True, exist_ok=True)

sys.path.insert(0, str(PIPELINE_DIR))

# ── Rich (pretty terminal) ────────────────────────────────────────────────────
try:
    from rich.console import Console
    from rich.panel import Panel
    from rich.table import Table
    from rich.rule import Rule
    from rich import box
    HAS_RICH = True
    console = Console()
except ImportError:
    HAS_RICH = False
    console = None

def _log(msg, style=""):
    if console:
        console.print(msg, style=style)
    else:
        print(msg)

def _ok(msg):   _log(f"  [green]✓[/green]  {msg}" if HAS_RICH else f"  ✓  {msg}")
def _warn(msg): _log(f"  [yellow]⚠[/yellow]  {msg}" if HAS_RICH else f"  ⚠  {msg}")
def _info(msg): _log(f"  [dim]{msg}[/dim]"           if HAS_RICH else f"     {msg}")
def _step(emoji, title):
    if HAS_RICH:
        console.print(f"\n[bold cyan]{emoji}  {title}[/bold cyan]")
        console.print(Rule(style="dim cyan"))
    else:
        print(f"\n{'─'*60}\n{emoji}  {title}")

# ── CSV sources (priority order) ──────────────────────────────────────────────
SOURCE_CSVS = [
    DATA_DIR / "linkedin_jobs.csv",
    DATA_DIR / "indeed_jobs.csv",
]

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _get(row, *keys, default=""):
    for k in keys:
        try:
            v = row.get(k)
            if v is not None and str(v).strip() not in ("", "nan", "None"):
                return str(v).strip()
        except Exception:
            pass
    return default

def _detect_platform(url: str) -> str:
    u = (url or "").lower()
    if "linkedin.com" in u:
        return "LinkedIn"
    if "indeed.com" in u:
        return "Indeed"
    return "External"


def _load_jobs(source_filter: str = "all") -> list[dict]:
    """
    Load all qualifying jobs from linkedin_jobs.csv + indeed_jobs.csv.
    Deduplicates by URL. Returns list of dicts.
    """
    try:
        import pandas as pd
    except ImportError:
        sys.exit("pip install pandas")

    all_jobs = []
    seen_urls = set()

    for csv_path in SOURCE_CSVS:
        if not csv_path.exists():
            _warn(f"  Not found (run scraper first): {csv_path.name}")
            continue

        try:
            df = pd.read_csv(csv_path)
        except Exception as e:
            _warn(f"  Could not read {csv_path.name}: {e}")
            continue

        if df.empty:
            continue

        url_col = next((c for c in df.columns if any(k in c.lower()
                       for k in ["apply_link", "job_url", "url", "link"])), None)
        if not url_col:
            _warn(f"  No URL column in {csv_path.name}")
            continue

        title_col = next((c for c in df.columns if "title" in c.lower()), None)
        co_col    = next((c for c in df.columns if any(k in c.lower()
                         for k in ["company", "employer", "organization"])), None)
        desc_col  = next((c for c in df.columns if any(k in c.lower()
                         for k in ["description", "desc", "detail"])), None)

        for _, row in df.iterrows():
            url      = _get(row, url_col, default="")
            title    = _get(row, title_col, default="") if title_col else ""
            company  = _get(row, co_col, default="")   if co_col    else ""
            jd       = _get(row, desc_col, default="") if desc_col  else ""
            platform = _detect_platform(url)

            if not url or not title or not company:
                continue
            if platform == "External":
                continue

            # Source filter
            if source_filter == "linkedin" and platform != "LinkedIn":
                continue
            if source_filter == "indeed" and platform != "Indeed":
                continue

            # Deduplicate
            key = url.split("?")[0].rstrip("/")
            if key in seen_urls:
                continue
            seen_urls.add(key)

            all_jobs.append({
                "title":    title,
                "company":  company,
                "url":      url,
                "jd":       jd,
                "platform": platform,
                "city":     _get(row, "job_city", "city", "location", default=""),
                "state":    _get(row, "job_state", "state", default=""),
            })

    return all_jobs


# ══════════════════════════════════════════════════════════════════════════════
# APPLY LOG
# ══════════════════════════════════════════════════════════════════════════════

def _load_log() -> list:
    if APPLY_LOG.exists():
        try:
            return json.loads(APPLY_LOG.read_text())
        except Exception:
            pass
    return []

def _save_log(entries: list):
    APPLY_LOG.write_text(json.dumps(entries, indent=2))

def _already_applied(url: str, log: list) -> bool:
    key = url.split("?")[0].rstrip("/")
    for entry in log:
        if entry.get("url", "").split("?")[0].rstrip("/") == key:
            if entry.get("status") in ("Applied", "Already Applied"):
                return True
    return False

def _log_entry(log: list, job: dict, status: str, note: str,
               fit_score: float, ats_before: float, ats_after: float,
               resume_path: str, screenshot: str = ""):
    log.append({
        "timestamp":   datetime.now().isoformat(),
        "company":     job["company"],
        "title":       job["title"],
        "platform":    job["platform"],
        "url":         job["url"],
        "fit_score":   fit_score,
        "ats_before":  ats_before,
        "ats_after":   ats_after,
        "status":      status,
        "note":        note,
        "resume_path": resume_path,
        "screenshot":  screenshot,
    })
    _save_log(log)


# ══════════════════════════════════════════════════════════════════════════════
# TRACKER  (Excel)
# ══════════════════════════════════════════════════════════════════════════════

def _update_tracker(log: list):
    """Rebuild Application_Tracker.xlsx from the apply log."""
    try:
        import pandas as pd
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        _warn("openpyxl not installed — tracker not updated")
        return

    if not log:
        return

    rows = []
    for e in log:
        rows.append({
            "Date":          e.get("timestamp", "")[:10],
            "Company":       e.get("company", ""),
            "Job Title":     e.get("title", ""),
            "Platform":      e.get("platform", ""),
            "Status":        e.get("status", ""),
            "Claude Fit %":  e.get("fit_score", 0),
            "ATS Before %":  e.get("ats_before", 0),
            "ATS After %":   e.get("ats_after", 0),
            "ATS Lift":      round(e.get("ats_after", 0) - e.get("ats_before", 0), 1),
            "Resume":        Path(e.get("resume_path", "")).name if e.get("resume_path") else "",
            "Note":          e.get("note", ""),
            "URL":           e.get("url", ""),
        })

    df = pd.DataFrame(rows)
    tracker_path = DATA_DIR / "Application_Tracker.xlsx"

    with pd.ExcelWriter(str(tracker_path), engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Applications")
        ws = writer.sheets["Applications"]

        # Header style
        header_fill = PatternFill("solid", fgColor="1F5C99")
        header_font = Font(bold=True, color="FFFFFF", name="Calibri")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Status colours
        status_colours = {
            "Applied":        "C6EFCE",
            "Already Applied":"BDD7EE",
            "Dry Run":        "FFEB9C",
            "Skipped":        "F2F2F2",
            "Failed":         "FFC7CE",
        }
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            status = str(row[4].value or "")
            fill_color = status_colours.get(status, "FFFFFF")
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(vertical="center")

        # Column widths
        widths = [12, 25, 35, 12, 14, 12, 12, 12, 10, 35, 30, 50]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

    _ok(f"Tracker updated → {tracker_path.name}  ({len(rows)} applications)")
    return tracker_path


# ══════════════════════════════════════════════════════════════════════════════
# PER-JOB PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

def process_job(job: dict, profile_summary: str, dry_run: bool,
                page_linkedin, page_indeed, log: list,
                ce, jdp, rb, cl_mod) -> dict:
    """
    Process ONE job end-to-end:
      score → parse JD → build resume → cover letter → apply → log
    Returns result dict.
    """
    title    = job["title"]
    company  = job["company"]
    url      = job["url"]
    jd       = job["jd"]
    platform = job["platform"]

    result = {
        "status":      "Skipped",
        "fit_score":   0,
        "ats_before":  0.0,
        "ats_after":   0.0,
        "resume_path": "",
        "note":        "",
    }

    # ── Step 1: Level filter ─────────────────────────────────────────────────
    if not ce.is_good_level(title):
        result["note"] = "Senior/Lead/Manager — filtered"
        return result

    # ── Step 2: Claude fit score ─────────────────────────────────────────────
    _info(f"  Scoring with Claude...")
    fit = ce.score_fit(profile_summary, jd, title, company)
    score     = float(fit.get("score", 0))
    grade     = fit.get("grade", "?")
    reasoning = fit.get("reasoning", "")
    result["fit_score"] = score

    score_bar = "█" * int(score // 10) + "░" * (10 - int(score // 10))
    _info(f"  [{score_bar}] {score:.0f}%  {grade}  — {reasoning[:80]}")

    if score < ce.FIT_THRESHOLD:
        result["note"] = f"Score {score:.0f}% below gate ({ce.FIT_THRESHOLD}%)"
        result["status"] = "Below Gate"
        return result

    # ── Step 3: Parse JD (ATS keywords + before/after score) ────────────────
    _info(f"  Parsing JD...")
    parsed = jdp.parse_jd(jd, title)
    ats_before = parsed.get("initial_score",   0.0)
    ats_after  = parsed.get("optimized_score", 0.0)
    jd_keywords    = parsed.get("jd_keywords",        [])
    injectable_kws = parsed.get("injectable_keywords", [])
    result["ats_before"] = ats_before
    result["ats_after"]  = ats_after
    _info(f"  ATS: {ats_before:.0f}% → {ats_after:.0f}%  ({len(jd_keywords)} JD keywords)")

    # ── Step 4: Build custom resume ──────────────────────────────────────────
    _info(f"  Building resume...")
    resume_path = ""
    try:
        res = rb.build_resume(
            job_title       = title,
            company         = company,
            jd_keywords     = jd_keywords,
            injectable_kws  = injectable_kws,
            initial_score   = ats_before,
            optimized_score = ats_after,
        )
        # build_resume returns (path, actual_initial, actual_optimized)
        if isinstance(res, tuple):
            resume_path = res[0]
            result["ats_before"] = res[1]  # use actual verified score
            result["ats_after"]  = res[2]
        else:
            resume_path = str(res)
        result["resume_path"] = resume_path
        _ok(f"  Resume → {Path(resume_path).name}")
    except Exception as e:
        _warn(f"  Resume build failed: {e}")
        result["note"] += f" | Resume error: {str(e)[:60]}"

    # ── Step 5: Cover letter ─────────────────────────────────────────────────
    _info(f"  Writing cover letter...")
    try:
        import raghav_profile as rp
        name = rp.PROFILE.get("name", "Raghavendra Karanam")
        cl_text = ce.write_cover_letter(name, profile_summary, jd, title, company)
        cl_path = cl_mod.save_cover_letter(cl_text, title, company)
        _ok(f"  Cover letter → {Path(cl_path).name}")
    except Exception as e:
        _warn(f"  Cover letter failed: {e}")

    # ── Step 6: Apply ────────────────────────────────────────────────────────
    apply_status = "Dry Run" if dry_run else "Pending"
    apply_note   = ""

    if dry_run:
        apply_status = "Dry Run"
        apply_note   = f"Would apply | Resume: {Path(resume_path).name if resume_path else 'none'}"
        _info(f"  [DRY RUN] Would apply to {company} ({platform})")
    else:
        job_for_apply = {**job, "resume": resume_path, "fit_score": score}
        screenshot = ""

        try:
            if platform == "LinkedIn" and page_linkedin:
                from auto_apply import apply_linkedin, take_screenshot
                apply_status, apply_note = apply_linkedin(page_linkedin, job_for_apply, dry_run=False)
                if apply_status == "Applied":
                    screenshot = take_screenshot(page_linkedin, f"{company}_{title}")

            elif platform == "Indeed" and page_indeed:
                from auto_apply import apply_indeed, take_screenshot
                apply_status, apply_note = apply_indeed(page_indeed, job_for_apply, dry_run=False)
                if apply_status == "Applied":
                    screenshot = take_screenshot(page_indeed, f"{company}_{title}")
            else:
                apply_status = "No Browser"
                apply_note   = f"No {platform} browser session available"

        except Exception as e:
            apply_status = "Error"
            apply_note   = str(e)[:120]

        icon = {"Applied":"✅","Already Applied":"🔵","Skipped":"⏭️",
                "Error":"❌","No Browser":"⚠️"}.get(apply_status, "•")
        _log(f"  {icon}  {apply_status}  — {apply_note}")

    result["status"] = apply_status
    result["note"]   = apply_note

    # ── Step 7: Log ──────────────────────────────────────────────────────────
    _log_entry(log, job, apply_status, apply_note,
               score, result["ats_before"], result["ats_after"],
               resume_path)

    return result


# ══════════════════════════════════════════════════════════════════════════════
# BROWSER SETUP
# ══════════════════════════════════════════════════════════════════════════════

LINKEDIN_SESSION = Path.home() / ".linkedin_session"
INDEED_SESSION   = Path.home() / ".indeed_session"

def _ensure_linkedin_session(pw):
    """Open LinkedIn browser. If not logged in, prompt user once."""
    LINKEDIN_SESSION.mkdir(parents=True, exist_ok=True)
    browser = pw.chromium.launch_persistent_context(
        user_data_dir=str(LINKEDIN_SESSION),
        headless=False,
        args=["--disable-blink-features=AutomationControlled"],
        viewport={"width": 1280, "height": 800},
    )
    page = browser.pages[0] if browser.pages else browser.new_page()
    try:
        page.goto("https://www.linkedin.com/feed/", wait_until="domcontentloaded", timeout=20000)
        time.sleep(2)
        if "feed" in page.url and page.locator("div.global-nav").count() > 0:
            _ok("LinkedIn: logged in (using saved session)")
            return browser, page
    except Exception:
        pass

    print()
    print("  🔐  LinkedIn login required.")
    print("      The browser opened — please log in to LinkedIn.")
    print("      After logging in, your session is SAVED PERMANENTLY.")
    print("      You will NEVER need to log in again.")
    print("      Press ENTER here once you can see your LinkedIn feed...")
    input("  → ")
    _ok("LinkedIn session saved → ~/.linkedin_session")
    return browser, page

def _ensure_indeed_session(pw):
    """Open Indeed browser. If not logged in, prompt user once."""
    INDEED_SESSION.mkdir(parents=True, exist_ok=True)
    browser = pw.chromium.launch_persistent_context(
        user_data_dir=str(INDEED_SESSION),
        headless=False,
        args=["--disable-blink-features=AutomationControlled"],
        viewport={"width": 1280, "height": 800},
    )
    page = browser.pages[0] if browser.pages else browser.new_page()
    try:
        page.goto("https://www.indeed.com/", wait_until="domcontentloaded", timeout=20000)
        time.sleep(2)
        if page.locator("[aria-label*='Account'], .gnav-profile-icon").count() > 0:
            _ok("Indeed: logged in (using saved session)")
            return browser, page
    except Exception:
        pass

    print()
    print("  🔐  Indeed login required.")
    print("      The browser opened — please log in to Indeed.")
    print("      After logging in, your session is SAVED PERMANENTLY.")
    print("      Press ENTER here once you're logged in...")
    input("  → ")
    _ok("Indeed session saved → ~/.indeed_session")
    return browser, page


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    p = argparse.ArgumentParser(description="Job Pipeline — Per-Job Sequential")
    p.add_argument("--dry-run",     action="store_true",
                   help="Score + build resumes + cover letters, but do NOT submit applications")
    p.add_argument("--no-apply",    action="store_true",
                   help="Build resumes only — don't open browser at all")
    p.add_argument("--skip-fetch",  action="store_true",
                   help="Use existing CSVs (this is the default — scraper runs first)")
    p.add_argument("--limit",       type=int, default=0,
                   help="Max jobs to process (0 = all)")
    p.add_argument("--source",      choices=["all", "linkedin", "indeed"], default="all",
                   help="Only process jobs from this platform")
    args = p.parse_args()

    dry_run  = args.dry_run
    no_apply = args.no_apply

    # ── Banner ────────────────────────────────────────────────────────────────
    if HAS_RICH:
        console.print()
        console.print(Panel.fit(
            "[bold cyan]🚀  JOB PIPELINE — Raghavendra Karanam[/bold cyan]\n"
            "[dim]LinkedIn Easy Apply  +  Indeed In-Portal  •  Per-Job Processing[/dim]\n"
            f"[dim]{datetime.now().strftime('%Y-%m-%d %H:%M')}[/dim]",
            border_style="cyan", padding=(1, 4),
        ))
        console.print()
    else:
        print("\n" + "="*60)
        print("  JOB PIPELINE — Per-Job Sequential  |  Raghavendra Karanam")
        print("="*60 + "\n")

    mode_str = "DRY RUN (no submissions)" if dry_run else ("BUILD ONLY (no browser)" if no_apply else "LIVE APPLY")
    _info(f"Mode   : {mode_str}")
    _info(f"Source : {args.source}")
    _info(f"Limit  : {args.limit if args.limit else 'all qualifying jobs'}")
    print()

    # ── Load modules ──────────────────────────────────────────────────────────
    try:
        import claude_engine  as ce
        import jd_parser      as jdp
        import resume_builder as rb
        import cover_letter   as cl_mod
        import raghav_profile as rp
    except ImportError as e:
        sys.exit(f"Import error: {e}\nRun: pip install anthropic rich pandas openpyxl python-docx playwright")

    # ── Build profile summary (once, reuse for all jobs) ─────────────────────
    full_profile = {
        **rp.PROFILE,
        "skills":     getattr(rp, "ALL_SKILLS_FLAT", []) or
                      [s for grp in getattr(rp, "SKILLS", {}).values() for s in grp],
        "experience": getattr(rp, "EXPERIENCE", []),
        "education":  getattr(rp, "EDUCATION",  []),
        "summary":    "Data professional with experience in data engineering, ETL, SQL, Python, Azure, Spark.",
    }
    profile_summary = ce.build_profile_summary(full_profile)
    _ok(f"Profile: {len(full_profile['skills'])} skills  |  {len(full_profile['experience'])} roles")

    # ── Load jobs ─────────────────────────────────────────────────────────────
    _step("📋", "Loading Jobs")
    jobs = _load_jobs(source_filter=args.source)

    if not jobs:
        _warn("No jobs found! Run linkedin_scraper.py or indeed_scraper.py first.")
        print()
        print("  From ~/job_pipeline, run:")
        print("    python linkedin_scraper.py")
        print("    python indeed_scraper.py")
        print()
        sys.exit(1)

    # Skip already-applied
    log = _load_log()
    applied_urls = {e.get("url", e.get("job_url", e.get("apply_link", ""))).split("?")[0].rstrip("/")
                    for e in log if e.get("status") in ("Applied", "Already Applied")}
    fresh = [j for j in jobs if j["url"].split("?")[0].rstrip("/") not in applied_urls]

    _ok(f"Total jobs in CSV  : {len(jobs)}")
    _ok(f"Already applied    : {len(jobs) - len(fresh)}")
    _ok(f"Ready to process   : {len(fresh)}")

    if args.limit > 0:
        fresh = fresh[:args.limit]
        _info(f"  (limited to {args.limit} by --limit flag)")

    if not fresh:
        _ok("Nothing to do — all jobs already applied to!")
        sys.exit(0)

    # ── Open browsers (unless --no-apply or --dry-run) ─────────────────────
    pw_ctx = None
    li_browser = li_page = None
    in_browser = in_page = None

    li_jobs  = [j for j in fresh if j["platform"] == "LinkedIn"]
    ind_jobs = [j for j in fresh if j["platform"] == "Indeed"]

    if not dry_run and not no_apply:
        try:
            from playwright.sync_api import sync_playwright
            pw_ctx = sync_playwright().__enter__()

            if li_jobs:
                _step("🔵", "LinkedIn Browser Setup")
                li_browser, li_page = _ensure_linkedin_session(pw_ctx)

            if ind_jobs:
                _step("🟡", "Indeed Browser Setup")
                in_browser, in_page = _ensure_indeed_session(pw_ctx)

        except ImportError:
            _warn("Playwright not installed — switching to --dry-run mode")
            dry_run = True

    # ── Per-job loop ──────────────────────────────────────────────────────────
    _step("⚡", f"Processing {len(fresh)} Jobs — One at a Time")

    counters = {
        "Applied": 0, "Dry Run": 0, "Already Applied": 0,
        "Skipped": 0, "Below Gate": 0, "Failed": 0, "Error": 0,
    }
    start = time.time()

    for i, job in enumerate(fresh, 1):
        platform_icon = "🔵" if job["platform"] == "LinkedIn" else "🟡"
        if HAS_RICH:
            console.print(f"\n[bold]{platform_icon} [{i}/{len(fresh)}]  "
                          f"{job['company'][:30]}  —  {job['title'][:40]}[/bold]")
            console.print(Rule(style="dim"))
        else:
            print(f"\n{'─'*60}")
            print(f"[{i}/{len(fresh)}] {job['platform']}  {job['company']} — {job['title']}")

        result = process_job(
            job            = job,
            profile_summary = profile_summary,
            dry_run        = dry_run or no_apply,
            page_linkedin  = li_page,
            page_indeed    = in_page,
            log            = log,
            ce             = ce,
            jdp            = jdp,
            rb             = rb,
            cl_mod         = cl_mod,
        )

        status = result.get("status", "Error")
        counters[status] = counters.get(status, 0) + 1

        # Update tracker after every job
        _update_tracker(log)

        # Polite pause between applications
        if status == "Applied" and not dry_run and not no_apply:
            time.sleep(4)

    # ── Close browsers ────────────────────────────────────────────────────────
    for browser in [li_browser, in_browser]:
        try:
            if browser:
                browser.close()
        except Exception:
            pass
    if pw_ctx:
        try:
            pw_ctx.__exit__(None, None, None)
        except Exception:
            pass

    # ── Final summary ─────────────────────────────────────────────────────────
    elapsed = time.time() - start
    m, s    = int(elapsed // 60), int(elapsed % 60)

    if HAS_RICH:
        t = Table(title="Pipeline Complete", box=box.ROUNDED, border_style="cyan")
        t.add_column("Result",  style="bold white", width=22)
        t.add_column("Count",   style="cyan",        width=8)
        for label, count in counters.items():
            if count > 0:
                icon = {"Applied":"✅","Dry Run":"👀","Below Gate":"⏭️",
                        "Skipped":"⏭️","Failed":"❌","Error":"❌"}.get(label,"•")
                t.add_row(f"{icon}  {label}", str(count))
        t.add_row("⏱  Time", f"{m}m {s}s")
        console.print(); console.print(t)

        tracker_path = DATA_DIR / "Application_Tracker.xlsx"
        console.print(Panel(
            f"[bold green]✅  Done![/bold green]\n\n"
            f"[dim]Resumes     → ~/job_pipeline/resumes/\n"
            f"Cover letters → ~/job_pipeline/cover_letters/\n"
            f"Tracker       → {tracker_path.name}\n"
            f"Log           → apply_log.json[/dim]",
            border_style="green", padding=(1, 2),
        ))
        console.print()
    else:
        print(f"\n{'='*50}")
        print(f"  Done in {m}m {s}s")
        for label, count in counters.items():
            if count > 0:
                print(f"  {label}: {count}")
        print('='*50)

    applied = counters.get("Applied", 0)
    if applied > 0:
        import subprocess
        tracker_path = DATA_DIR / "Application_Tracker.xlsx"
        if tracker_path.exists():
            subprocess.Popen(["open", str(tracker_path)])


if __name__ == "__main__":
    main()
