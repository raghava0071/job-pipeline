# =============================================================================
# TRACKER.PY — PRO v6  (Claude Fit Score + Funnel Tracking)
#
# COLUMNS:
#   A  #      B  Date Applied    C  Company         D  Job Title
#   E  Location     F  Claude Score   G  Grade          H  Status
#   I  Apply Link   J  Follow-Up      K  Fit Reasoning  L  Notes
#   M  Resume File  N  Cover Letter   O  Salary Range
#   P  ATS Response     Q  Response Days   R  Interview Stage   S  Rejection Reason
#
# Columns P–S are the FUNNEL TRACKER. Update these manually as responses come in:
#   P: ATS Response  — "Auto-Reject", "No Response", "Callback", "Pending"
#   Q: Response Days — how many days from apply date to first response
#   R: Interview Stage — "None", "Phone Screen", "Technical", "Final Round", "Offer"
#   S: Rejection Reason — "ATS filtered", "Overqualified", "No sponsorship", "Ghosted", etc.
#
# Use the Dashboard tab to see funnel drop-off rates (where you're losing candidates)
# =============================================================================

from __future__ import annotations

import os
import sys
import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl --break-system-packages")
    sys.exit(1)

TRACKER_PATH = Path(__file__).parent / "data" / "Application_Tracker.xlsx"

# ── palette ───────────────────────────────────────────────────────────────────
BLUE_HEADER = "1F5C99"
WHITE       = "FFFFFF"
GRAY_LIGHT  = "F2F2F2"
DARK_TEXT   = "222222"

GRADE_COLORS = {
    "A":  "C6EFCE",   # green  — 85-100
    "B+": "DDEEFF",   # light blue — 77-84
    "B":  "DDEEFF",   # light blue — 70-76
    "C":  "FFEB9C",   # yellow — 55-69
    "D":  "FCE4D6",   # red    — <55
}

STATUS_COLORS = {
    "Applied":             "FFF2CC",
    "Needs Manual Apply":  "FFE0B2",
    "Already Applied":     "E3F2FD",
    "Phone Screen":        "DDEBF7",
    "Interview":           "E2EFDA",
    "Offer":               "C6EFCE",
    "Rejected":            "FFCCCC",
    "No Response":         "F2F2F2",
    "Withdrawn":           "EDEDED",
}

# ── column definitions ────────────────────────────────────────────────────────
COLUMNS = [
    ("A",  "#",                5),
    ("B",  "Date Applied",    14),
    ("C",  "Company",         24),
    ("D",  "Job Title",       30),
    ("E",  "Location",        18),
    ("F",  "Claude Score",    13),
    ("G",  "Grade",            8),
    ("H",  "Status",          20),
    ("I",  "Apply Link",      28),
    ("J",  "Follow-Up",       14),
    ("K",  "Fit Reasoning",   40),
    ("L",  "Notes",           30),
    ("M",  "Resume File",     32),
    ("N",  "Cover Letter",    32),
    ("O",  "Salary Range",    18),
    # ── Funnel tracking columns — fill these in as responses arrive ──────────
    ("P",  "ATS Response",    18),   # Auto-Reject | No Response | Callback | Pending
    ("Q",  "Response Days",   14),   # Days from apply to first response (number)
    ("R",  "Interview Stage", 18),   # None | Phone Screen | Technical | Final | Offer
    ("S",  "Rejection Reason",28),   # ATS filtered | Ghosted | Overqualified | No sponsorship
]

COL_SCORE  = "F"
COL_GRADE  = "G"
COL_STATUS = "H"
COL_LINK   = "I"
COL_ATS    = "P"
COL_STAGE  = "R"


# ── style helpers ─────────────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _border(color: str = "CCCCCC") -> Border:
    side = Side(style="hair", color=color)
    return Border(bottom=side, right=side)

def _header_style(ws) -> None:
    for col_letter, col_name, col_width in COLUMNS:
        cell = ws[f"{col_letter}1"]
        cell.value     = col_name
        cell.font      = Font(name="Calibri", bold=True, color=WHITE, size=10)
        cell.fill      = _fill(BLUE_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = Border(
            bottom=Side(style="medium", color=WHITE),
            right= Side(style="thin",   color="4472C4"),
        )
        ws.column_dimensions[col_letter].width = col_width
    ws.row_dimensions[1].height = 36

def _row_style(ws, row: int, status: str, grade: str = "") -> None:
    # Row background = status color
    bg = STATUS_COLORS.get(status, WHITE)
    for col_letter, _, _ in COLUMNS:
        cell = ws[f"{col_letter}{row}"]
        cell.fill      = _fill(bg)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border    = _border()
        cell.font      = Font(name="Calibri", size=10, color=DARK_TEXT)
    # Override grade cell color
    if grade and grade in GRADE_COLORS:
        ws[f"{COL_GRADE}{row}"].fill = _fill(GRADE_COLORS[grade])
        ws[f"{COL_GRADE}{row}"].font = Font(name="Calibri", bold=True, size=11,
                                            color=DARK_TEXT)
        ws[f"{COL_GRADE}{row}"].alignment = Alignment(horizontal="center",
                                                      vertical="center")
    ws.row_dimensions[row].height = 24


def _apply_cf(ws, num_rows: int) -> None:
    if num_rows < 1:
        return
    last = num_rows + 1
    # Claude Score color coding: <55=red, 55-69=yellow, 70-84=blue, >=85=green
    ws.conditional_formatting.add(f"F2:F{last}",
        CellIsRule("lessThan", ["55"], fill=_fill("FCE4D6")))
    ws.conditional_formatting.add(f"F2:F{last}",
        CellIsRule("between", ["55", "69"], fill=_fill("FFEB9C")))
    ws.conditional_formatting.add(f"F2:F{last}",
        CellIsRule("between", ["70", "84"], fill=_fill("DDEEFF")))
    ws.conditional_formatting.add(f"F2:F{last}",
        CellIsRule("greaterThanOrEqual", ["85"], fill=_fill("C6EFCE")))


# ── column value helpers ───────────────────────────────────────────────────────
def _get(row, *keys, default=""):
    """Try multiple column name keys, return first non-empty value."""
    for k in keys:
        try:
            v = row.get(k)
            if v is not None and str(v).strip() not in ("", "nan", "None"):
                return str(v).strip()
        except Exception:
            continue
    return default

def _safe_float(row, *keys, default=0.0):
    for k in keys:
        try:
            v = row.get(k)
            if v is not None and str(v).strip() not in ("", "nan", "None"):
                return float(v)
        except Exception:
            continue
    return default

def _salary_range(row):
    lo = _safe_float(row, "job_min_salary", "min_salary")
    hi = _safe_float(row, "job_max_salary", "max_salary")
    curr = _get(row, "job_salary_currency", "salary_currency", default="USD")
    if lo and hi:
        return f"${lo:,.0f}–${hi:,.0f}"
    elif hi:
        return f"Up to ${hi:,.0f}"
    elif lo:
        return f"${lo:,.0f}+"
    return _get(row, "salary", default="TBD")

def _location(row):
    city  = _get(row, "job_city",   "city")
    state = _get(row, "job_state",  "state")
    remote = _get(row, "job_is_remote", "is_remote")
    if str(remote).lower() in ("true", "1", "yes", "remote"):
        return f"Remote{f', {state}' if state else ''}"
    parts = [p for p in [city, state] if p]
    return ", ".join(parts) if parts else _get(row, "location", default="")


# =============================================================================
# MAIN TRACKER BUILDER
# =============================================================================

def create_tracker(jobs_df: pd.DataFrame = None) -> str:
    TRACKER_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # ── Sheet 1: Applications ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Applications"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    _header_style(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    today      = datetime.now()
    rows_added = 0
    scores: list[float] = []
    grade_counts = {"A": 0, "B": 0, "C": 0, "D": 0}

    if jobs_df is not None and not jobs_df.empty:
        for i, (_, row) in enumerate(jobs_df.iterrows(), start=2):
            follow_up = (today + timedelta(days=7)).strftime("%Y-%m-%d")

            # ── Core fields (correct column names from pipeline) ───────────────
            company   = _get(row, "employer_name", "company", "employer")
            title_val = _get(row, "job_title",     "title",   "job_position")
            location  = _location(row)
            apply_link= _get(row, "job_apply_link","apply_link","url")
            salary    = _salary_range(row)

            # ── Claude fit fields ─────────────────────────────────────────────
            fit_score   = _safe_float(row, "fit_score",   default=0.0)
            fit_grade   = _get(row, "fit_grade",   default="—")
            fit_reason  = _get(row, "fit_reasoning", default="")
            fit_apply   = str(row.get("fit_apply", "")).lower() in ("true", "1", "yes")

            scores.append(fit_score)
            if fit_grade in grade_counts:
                grade_counts[fit_grade] += 1

            # ── Status ────────────────────────────────────────────────────────
            status = _get(row, "status", default="Needs Manual Apply")
            if status in ("nan", "None", "Pending", ""):
                status = "Needs Manual Apply"

            # ── Resume / cover letter file names ──────────────────────────────
            resume_path = _get(row, "resume_path",       default="")
            cl_path     = _get(row, "cover_letter_path", default="")
            if not resume_path:
                safe_co = company[:20].replace(" ", "_")
                safe_ti = title_val[:15].replace(" ", "_")
                resume_path = f"Raghavendra_Karanam_{safe_co}_{safe_ti}.docx"
            if not cl_path:
                safe_co = company[:20].replace(" ", "_")
                safe_ti = title_val[:15].replace(" ", "_")
                cl_path = f"CoverLetter_{safe_co}_{safe_ti}.docx"

            # ── Build row data ────────────────────────────────────────────────
            data = [
                i - 1,                            # A: #
                today.strftime("%Y-%m-%d"),        # B: Date
                company,                           # C: Company
                title_val,                         # D: Job Title
                location,                          # E: Location
                round(fit_score, 0),               # F: Claude Score
                fit_grade,                         # G: Grade
                status,                            # H: Status
                apply_link,                        # I: Apply Link
                follow_up,                         # J: Follow-Up
                fit_reason[:200],                  # K: Fit Reasoning
                _get(row, "apply_note", default=""),# L: Notes
                os.path.basename(resume_path),     # M: Resume File
                os.path.basename(cl_path),         # N: Cover Letter
                salary,                            # O: Salary
                _get(row, "ats_response", default="Pending"),   # P: ATS Response
                _get(row, "response_days", default=""),         # Q: Response Days
                _get(row, "interview_stage", default="None"),   # R: Interview Stage
                _get(row, "rejection_reason", default=""),      # S: Rejection Reason
            ]

            for col_idx, value in enumerate(data, start=1):
                cl = get_column_letter(col_idx)
                cell = ws[f"{cl}{i}"]
                cell.value = value

                # Hyperlink for apply link
                if cl == COL_LINK and str(value).startswith("http"):
                    cell.hyperlink = str(value)
                    cell.value     = "Apply →"
                    cell.font      = Font(name="Calibri", size=10,
                                         color=BLUE_HEADER, underline="single")

                # Claude score: center + number format
                if cl == COL_SCORE:
                    cell.alignment    = Alignment(horizontal="center", vertical="center")
                    cell.number_format = '0"%"'
                    cell.font         = Font(name="Calibri", bold=True, size=11,
                                            color=DARK_TEXT)

                # Status center
                if cl == COL_STATUS:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            _row_style(ws, i, status, fit_grade)
            rows_added += 1

    else:
        # Placeholder row
        sample = [
            1, today.strftime("%Y-%m-%d"), "Acme Analytics", "Data Analyst",
            "Remote, TX", 78, "B+", "Needs Manual Apply",
            "https://linkedin.com/jobs/view/123",
            (today + timedelta(days=7)).strftime("%Y-%m-%d"),
            "Strong SQL and Python skills match well. Missing 2 years experience.",
            "", "Raghavendra_Karanam_Acme_Data_Analyst.docx",
            "CoverLetter_Acme_Data_Analyst.docx", "$90,000–$110,000",
        ]
        for col_idx, value in enumerate(sample, start=1):
            cell = ws[f"{get_column_letter(col_idx)}2"]
            cell.value = value
        _row_style(ws, 2, "Needs Manual Apply", "B+")
        rows_added = 1

    _apply_cf(ws, rows_added)

    # ── Sheet 2: Dashboard ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Dashboard")
    ws2.sheet_view.showGridLines = False
    for col, w in [("A", 32), ("B", 18), ("D", 32), ("F", 22)]:
        ws2.column_dimensions[col].width = w

    ws2["A1"] = "APPLICATION DASHBOARD — Claude-Powered Pipeline"
    ws2["A1"].font      = Font(name="Calibri", bold=True, size=16, color=BLUE_HEADER)
    ws2["A2"] = f"Raghavendra Karanam  |  Data Analyst · Engineer · Scientist  |  {today:%b %d, %Y}"
    ws2["A2"].font      = Font(name="Calibri", size=10, color="888888")
    ws2.row_dimensions[1].height = 28

    avg_score = sum(scores) / len(scores) if scores else 0.0
    good_fits = sum(1 for s in scores if s >= 65)
    total = rows_added

    # Count statuses from the same logic used when writing rows
    # (don't rely on "status" column existing in df — it usually won't)
    status_counts: dict[str, int] = {}
    if jobs_df is not None and not jobs_df.empty:
        for _, row in jobs_df.iterrows():
            s = _get(row, "status", default="Needs Manual Apply")
            if s in ("nan", "None", "Pending", ""):
                s = "Needs Manual Apply"
            status_counts[s] = status_counts.get(s, 0) + 1

    # Compute funnel counts from df if available
    ats_response_counts:  dict[str, int] = {}
    interview_stage_counts: dict[str, int] = {}
    rejection_reason_counts: dict[str, int] = {}
    if jobs_df is not None and not jobs_df.empty:
        for _, row in jobs_df.iterrows():
            ats = _get(row, "ats_response", default="Pending")
            ats_response_counts[ats] = ats_response_counts.get(ats, 0) + 1
            stage = _get(row, "interview_stage", default="None")
            interview_stage_counts[stage] = interview_stage_counts.get(stage, 0) + 1
            rej = _get(row, "rejection_reason", default="")
            if rej:
                rejection_reason_counts[rej] = rejection_reason_counts.get(rej, 0) + 1

    callback_rate = 0.0
    if total > 0:
        callbacks = ats_response_counts.get("Callback", 0)
        callback_rate = callbacks / total * 100

    stats = [
        ("── PIPELINE STATS ──",                    ""),
        ("Total Applications",                       total),
        ("Good Fits (Claude ≥ 72%)",                 good_fits),
        ("Avg Claude Fit Score",                     f"{avg_score:.1f}%"),
        ("",                                         ""),
        ("── GRADE BREAKDOWN ──",                    ""),
        ("A  Grade  (85–100%)",                      grade_counts.get("A", 0)),
        ("B  Grade  (70–84%)",                       grade_counts.get("B", 0)),
        ("C  Grade  (55–69%)",                       grade_counts.get("C", 0)),
        ("D  Grade  (<55%)  — skipped",              grade_counts.get("D", 0)),
        ("",                                         ""),
        ("── APPLICATION STATUS ──",                 ""),
        ("Applied (Easy Apply / Submitted)",         status_counts.get("Applied", 0)),
        ("Needs Manual Apply",                       status_counts.get("Needs Manual Apply", 0)),
        ("Phone Screen",                             status_counts.get("Phone Screen", 0)),
        ("Interview",                                status_counts.get("Interview", 0)),
        ("Offer",                                    status_counts.get("Offer", 0)),
        ("Rejected",                                 status_counts.get("Rejected", 0)),
        ("",                                         ""),
        ("── FUNNEL (update manually) ──",           ""),
        ("ATS Auto-Rejected",                        ats_response_counts.get("Auto-Reject", 0)),
        ("No Response (ghosted)",                    ats_response_counts.get("No Response", 0)),
        ("Callbacks received",                       ats_response_counts.get("Callback", 0)),
        ("Callback rate",                            f"{callback_rate:.1f}%"),
        ("Reached Phone Screen",                     interview_stage_counts.get("Phone Screen", 0)),
        ("Reached Technical Round",                  interview_stage_counts.get("Technical", 0)),
        ("Reached Final Round",                      interview_stage_counts.get("Final Round", 0)),
        ("",                                         ""),
        ("── WHERE TO LOOK IF 0 CALLBACKS ──",       ""),
        ("• Update col P with email responses",      "Auto-Reject vs No Response"),
        ("• If mostly Auto-Reject → ATS issue",      "Resume keywords need work"),
        ("• If mostly No Response → Recruiter issue","Resume quality / fit score too low"),
        ("• If Callbacks but no screens → Phone prep","Practice intro + elevator pitch"),
        ("",                                         ""),
        ("Last Updated",                             today.strftime("%Y-%m-%d %H:%M")),
    ]

    row_num = 4
    for label, value in stats:
        if not label:
            row_num += 1
            continue
        cl = ws2[f"A{row_num}"]
        cv = ws2[f"B{row_num}"]
        cl.value = label
        cv.value = value
        is_header = label.startswith("──")
        cl.font   = Font(name="Calibri", bold=True, size=11,
                         color=BLUE_HEADER if is_header else DARK_TEXT)
        cv.font   = Font(name="Calibri", size=11, bold=is_header,
                         color=("00B050" if any(k in label for k in
                                ["Good Fits", "Applied", "Offer", "Avg"]) else BLUE_HEADER))
        bg = "E8F0F9" if is_header else (GRAY_LIGHT if row_num % 2 == 0 else WHITE)
        cl.fill = _fill(bg)
        cv.fill = _fill(bg)
        row_num += 1

    # ── Score color guide ─────────────────────────────────────────────────────
    ws2["D4"] = "Claude Score Guide"
    ws2["D4"].font = Font(name="Calibri", bold=True, size=11, color=BLUE_HEADER)
    guide = [
        ("A  ≥ 85%   → Excellent fit",  "C6EFCE"),
        ("B  70–84%  → Good fit",       "DDEEFF"),
        ("C  55–69%  → Moderate fit",   "FFEB9C"),
        ("D  < 55%   → Poor fit",       "FCE4D6"),
        ("",                            WHITE),
        ("Green  → Apply queue",        "C6EFCE"),
        ("Orange → Manual apply",       "FFE0B2"),
        ("Yellow → Applied",            "FFF2CC"),
        ("Blue   → Phone/Interview",    "DDEBF7"),
        ("Pink   → Rejected",           "FFCCCC"),
    ]
    for j, (lbl, color) in enumerate(guide, start=5):
        c = ws2[f"D{j}"]
        c.value = lbl
        c.fill  = _fill(color)
        c.font  = Font(name="Calibri", size=10)
        c.alignment = Alignment(horizontal="left", indent=1)
        if color != WHITE:
            ws2.row_dimensions[j + 3].height = 20

    wb.save(str(TRACKER_PATH))

    print(f"\n  ✅  Tracker saved → {TRACKER_PATH}")
    print(f"  Rows: {total}  |  Good fits: {good_fits}  |  Avg Claude score: {avg_score:.1f}%")
    print(f"  Grades: A={grade_counts['A']}  B={grade_counts['B']}  "
          f"C={grade_counts['C']}  D={grade_counts['D']}")
    return str(TRACKER_PATH)


# =============================================================================
# STATUS UPDATER
# =============================================================================

def update_status(company_or_id: str, new_status: str, notes: str = "") -> None:
    if not TRACKER_PATH.exists():
        return
    wb = openpyxl.load_workbook(str(TRACKER_PATH))
    ws = wb["Applications"]
    COL_IDX = {c: i for i, (c, _, _) in enumerate(COLUMNS, start=1)}
    status_col  = COL_IDX["H"]
    company_col = COL_IDX["C"]
    notes_col   = COL_IDX["L"]
    for row in ws.iter_rows(min_row=2):
        company_val = str(row[company_col - 1].value or "").lower()
        if company_or_id.lower() in company_val:
            row[status_col - 1].value = new_status
            if notes:
                existing = str(row[notes_col - 1].value or "")
                row[notes_col - 1].value = (f"{existing}\n{notes}".strip()
                                            if existing else notes)
            bg = STATUS_COLORS.get(new_status, WHITE)
            for cell in row:
                if cell.value is not None:
                    cell.fill = PatternFill("solid", fgColor=bg)
            break
    wb.save(str(TRACKER_PATH))


# =============================================================================
# STANDALONE
# =============================================================================
if __name__ == "__main__":
    filtered_csv = Path(__file__).parent / "data" / "filtered_jobs.csv"
    if filtered_csv.exists():
        df = pd.read_csv(filtered_csv)
        create_tracker(df)
    else:
        print("No filtered_jobs.csv — creating tracker with placeholder row")
        create_tracker()
